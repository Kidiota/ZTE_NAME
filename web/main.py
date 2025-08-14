import pdfplumber
import PyPDF2
import os
import shutil
import xlrd
import time

#1111

def fix_cropbox(pdf_path, output_path):
    """修复PDF的CropBox并写入output_path"""
    with open(pdf_path, "rb") as file:
        reader = PyPDF2.PdfReader(file)
        writer = PyPDF2.PdfWriter()
        for page in reader.pages:
            try:
                if "/CropBox" not in page:
                    page.cropbox = page.mediabox
            except Exception:
                try:
                    page.cropbox = page.mediabox
                except Exception:
                    pass
            writer.add_page(page)
    with open(output_path, "wb") as output_file:
        writer.write(output_file)

def get_raw_info(filesName, fixed_folder):
    """从修复后的PDF中提取每行文本"""
    fixedFileFullName = os.path.join(fixed_folder, filesName)
    lines = []
    with pdfplumber.open(fixedFileFullName) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                for line in text.splitlines():
                    lines.append(line + "\n")
    return lines

def make_it_readable(raw_data):
    """直接返回行列表"""
    return raw_data if raw_data else []

def get_pdf_info(pdf_data):
    """从 pdf_data 中提取 PO Number 和 FINAL/COA 信息（大小写无关）"""
    pdf_info = []
    finalOrNot = ""
    poNo = ""
    for line in pdf_data:
        # 忽略大小写和前后空格
        text = line.strip().upper()

        if text == "FINAL CERTIFICATE OF ACCEPTANCE":
            finalOrNot = "FCOA"
        elif text == "CERTIFICATE OF ACCEPTANCE":
            finalOrNot = "COA"

        if text.startswith("PO NUMBER :") or text.startswith("PO NUMBER:"):
            remainder = line.split(":", 1)[1].strip()
            poNo_candidate = remainder.split()[0]
            poNo = poNo_candidate
            pdf_info = [poNo, finalOrNot]
            break
    return pdf_info

def read_xlsx(xlsxFileName):
    """读取 xlsx/xls 文件"""
    try:
        xlsxData = xlrd.open_workbook(xlsxFileName).sheets()[0]
        return xlsxData
    except Exception:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsxFileName, read_only=True, data_only=True)
            sheet = wb[wb.sheetnames[0]]
            class SheetWrapper:
                def __init__(self, sheet):
                    self._rows = list(sheet.values)
                def col_values(self, idx):
                    return [row[idx] if idx < len(row) else None for row in self._rows]
                def cell_value(self, r, c):
                    return self._rows[r][c] if c < len(self._rows[r]) else None
            return SheetWrapper(sheet)
        except Exception as e:
            raise RuntimeError(f"无法读取 xlsx 文件: {e}")

def get_data_from_xlsx(poNo, xlsxData):
    """根据 PO Number 从 xlsx 表格查找 siteID 和 projectName"""
    try:
        search_key = float(poNo)
        col0 = xlsxData.col_values(0)
        loPO = col0.index(search_key)
        siteID = xlsxData.cell_value(loPO, 23)
        projectName = xlsxData.cell_value(loPO, 1)
        return [siteID, projectName]
    except Exception:
        try:
            col0 = xlsxData.col_values(0)
            for idx, val in enumerate(col0):
                if val and str(val).strip() == str(poNo).strip():
                    siteID = xlsxData.cell_value(idx, 23)
                    projectName = xlsxData.cell_value(idx, 1)
                    return [siteID, projectName]
        except Exception:
            pass
    return None

def _sanitize_filename(name):
    """清理非法文件名字符"""
    import re
    name = str(name).strip()
    name = re.sub(r'[\\/*?:"<>|]', '_', name)
    return name.replace(' ', '_')

def process_files(input_folder, xlsx_path, output_folder):
    """处理 PDF 与 XLSX 文件"""
    fixed_folder = os.path.join(input_folder, "fixed")
    if os.path.exists(fixed_folder):
        shutil.rmtree(fixed_folder)
    os.mkdir(fixed_folder)

    error_files = []

    # 1) 修复 PDF
    for file_name in os.listdir(input_folder):
        if not file_name.lower().endswith(".pdf"):
            continue
        src_path = os.path.join(input_folder, file_name)
        dst_path = os.path.join(fixed_folder, "fixed_" + file_name)
        try:
            fix_cropbox(src_path, dst_path)
        except Exception:
            try:
                shutil.copyfile(src_path, dst_path)
            except Exception:
                error_files.append(file_name)
                continue

    # 2) 读取 xlsx
    xlsxData = read_xlsx(xlsx_path)

    # 3) 处理 PDF
    for fixed_name in os.listdir(fixed_folder):
        fixed_path = os.path.join(fixed_folder, fixed_name)
        original_name = fixed_name[6:] if fixed_name.startswith("fixed_") else fixed_name
        try:
            raw_lines = get_raw_info(fixed_name, fixed_folder)
            if not raw_lines:
                out_name = "unmatched_" + _sanitize_filename(original_name)
                shutil.copyfile(fixed_path, os.path.join(output_folder, out_name))
                error_files.append(original_name)
                continue

            pdf_data = make_it_readable(raw_lines)
            pdf_info = get_pdf_info(pdf_data)
            if not pdf_info or not pdf_info[0]:
                out_name = "unmatched_" + _sanitize_filename(original_name)
                shutil.copyfile(fixed_path, os.path.join(output_folder, out_name))
                error_files.append(original_name)
                continue

            poNo = pdf_info[0]
            finalOrNot = pdf_info[1] if len(pdf_info) > 1 else ""

            xlsx_info = get_data_from_xlsx(poNo, xlsxData)
            if not xlsx_info:
                out_name = "unmatched_PO" + _sanitize_filename(original_name)
                shutil.copyfile(fixed_path, os.path.join(output_folder, out_name))
                error_files.append(original_name)
                continue

            siteID, projectName = xlsx_info
            timeNow = time.strftime('%Y-%m-%d', time.localtime())
            out_filename = f"{_sanitize_filename(siteID)}_TM_{_sanitize_filename(projectName)}_PO{_sanitize_filename(poNo)}_{_sanitize_filename(finalOrNot)}_{timeNow}.pdf"
            shutil.copyfile(fixed_path, os.path.join(output_folder, out_filename))
        except Exception:
            try:
                out_name = "error_" + _sanitize_filename(original_name)
                shutil.copyfile(fixed_path, os.path.join(output_folder, out_name))
            except Exception:
                pass
            error_files.append(original_name)

    return error_files
